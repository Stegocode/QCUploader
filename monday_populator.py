"""
monday_populator.py
===================
Scrapes the batch invoice from the order management platform,
parses orders, and creates missing items in the Monday delivery board.
"""

import json
import os
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

load_dotenv()

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_URL     = os.getenv("PLATFORM_BASE_URL", "https://your-platform.example.com")
DOWNLOAD_DIR = Path(os.getenv("DOWNLOAD_DIR", Path.home() / "order_photo_sync" / "inbox"))

MONDAY_API_URL      = "https://api.monday.com/v2"
TRANSFERRED_GROUP_ID = os.getenv("MONDAY_TRANSFERRED_GROUP_ID", "")

EXCLUDED_TRUCKS = {"STORAGE", "UNPAID"}

# ---------------------------------------------------------------------------
# Product classification sets
# ---------------------------------------------------------------------------

DELETE_CODES = {
    "B003", "B002", "CC FEES", "CC FEE", "STAIR 5-10", "STAIR 11-15",
    "STAIR 16-20", "STAIR 21-25", "ACCOMM", "MGMT ACCOMMODATION",
    "R&O-ACCTG USE ONLY", "PENCE - ACCTG USE ONLY", "WALSH -ACCTG USE ONLY",
    "COLAS-ACCTG USE ONLY", "LIEN", "CAT TAX", "NSF", "CHARGEBACK CC",
    "PAYROLL", "SM WO", "ACCTG - WARRANTIES", "ACCOUNTING", "REFUND",
    "WF FEES", "TD", "FACTORY", "CC FEE", "STORAGE", "RESTOCK", "LATE FEE",
    "SPECIAL", "99PRICEADJ", "99MISC50", "CUSTOMERS HOME",
    "DAMAGE BY DELIVERY", "FREIGHT",
}

SERVICE_CODES = {
    "X001", "X002", "X003", "X004", "X007", "X011", "X012", "X013", "X014",
    "X015", "X018", "X020", "X021", "X022", "X023", "X025", "X028", "X029",
    "X075", "X100", "X101", "X102", "X103", "X103A", "X103B", "X151", "X152",
    "X154", "X156", "X157", "X159", "X164", "X166", "X201", "X204", "X208",
    "X210", "X212", "X214", "X216", "X222", "X224", "X226", "X230", "X232",
    "X238", "X251", "X253", "X255", "X257", "X259", "X260", "X261", "X301",
    "X303", "X328", "X336", "X998", "MEMO",
}

PARTS_LOOKUP = {
    "WATERLINE", "WATERHOSE SS", "WATERHOSE RUB", "GASLINE", "LAUNDRYPACK-GAS",
    "DRYERCORD", "DWELBOW", "BRACKET", 'ADA DW PANS 18"', "DW PANS",
    "LAUNDRY PAN", "52525", "3PRONGCORD", "110 POWERCORD", "DW KIT",
    "LAUNDRYPACK", "LAUNDRYPACK-ELECT", "RANGECORD", "DW INSTALL KIT",
    "STEAM DRYER", "STEAM DRYER ",
}

LABOR_LOOKUP        = {"WASHERIN", "REFERINSTALL", "CONVERSION-DOORSWING"}
LABOR_PART_CODES    = {"TPI", "CONVERSION-GAS"}
DRYER_INSTALL_CODES = {"DRYER INSTALL"}
REDEL_CODES         = {"REDEL"}

NON_MODEL_CODES = (
    DELETE_CODES | SERVICE_CODES | PARTS_LOOKUP |
    LABOR_LOOKUP | LABOR_PART_CODES | DRYER_INSTALL_CODES | REDEL_CODES
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def is_model(model_key: str) -> bool:
    """Return True if the line item represents a physical product model."""
    return bool(model_key) and model_key not in NON_MODEL_CODES


def get_next_business_day() -> datetime:
    today = datetime.today()
    delta = 3 if today.weekday() == 4 else (2 if today.weekday() == 5 else 1)
    return today + timedelta(days=delta)


def _monday_headers(token: str) -> dict:
    return {"Authorization": token, "Content-Type": "application/json"}


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------

def scrape_batch_invoice(hs_username: str, hs_password: str, delivery_date_str: str) -> str | None:
    """
    Log into the platform and download the batch invoice Excel file
    for the given delivery date. Returns the local file path or None.
    """
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    for f in DOWNLOAD_DIR.iterdir():
        try:
            f.unlink()
        except Exception:
            pass

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_experimental_option("prefs", {
        "download.default_directory": str(DOWNLOAD_DIR),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    })

    driver = webdriver.Chrome(options=options)
    wait   = WebDriverWait(driver, 40)

    def js_click(el):
        driver.execute_script("arguments[0].click();", el)

    def wait_for_download(keyword: str, timeout: int = 30) -> str | None:
        print(f"  Waiting for download: {keyword}")
        for _ in range(timeout):
            matches = [
                f for f in DOWNLOAD_DIR.iterdir()
                if keyword in f.name.lower() and not f.name.endswith(".crdownload")
            ]
            if matches:
                print(f"  Downloaded: {matches[0].name}")
                return str(matches[0])
            time.sleep(1)
        print(f"  Timed out waiting for: {keyword}")
        return None

    try:
        print("  Logging in...")
        driver.get(f"{BASE_URL}/login")
        time.sleep(8)
        wait.until(EC.presence_of_element_located((By.NAME, "email")))
        driver.find_element(By.NAME, "email").send_keys(hs_username)
        driver.find_element(By.NAME, "password").send_keys(hs_password)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(4)
        print("  Logged in")

        try:
            save_btn = wait.until(EC.presence_of_element_located((By.ID, "save-current-location")))
            time.sleep(1)
            js_click(save_btn)
            time.sleep(3)
            print("  Location saved")
        except Exception:
            print("  No location popup")

        print("  Navigating to batch invoice...")
        driver.get(f"{BASE_URL}/sales/batch-invoice")
        time.sleep(5)

        print(f"  Setting date to {delivery_date_str}")
        date_field = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "input.form-control.input[type='text']")
        ))
        driver.execute_script("arguments[0].value = '';", date_field)
        date_field.click()
        time.sleep(0.5)
        date_field.send_keys(Keys.CONTROL + "a")
        date_field.send_keys(Keys.DELETE)
        time.sleep(0.3)
        date_field.send_keys(delivery_date_str)
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", date_field
        )
        time.sleep(1)
        date_field.send_keys(Keys.RETURN)
        time.sleep(5)

        print("  Clicking export...")
        export_btn = driver.find_element(By.XPATH, "//button[@onclick='batchPrintExcel()']")
        js_click(export_btn)
        batch_file = wait_for_download("bulk-invoice", timeout=30) or wait_for_download(".xlsx", timeout=15)
        return batch_file

    except Exception as e:
        print(f"  Scrape error: {e}")
        return None
    finally:
        driver.quit()


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def parse_orders(batch_file: str) -> dict:
    """
    Parse the batch invoice Excel file and return a dict of
    {order_number: row_count} representing items to create in Monday.
    """
    print(f"  Parsing: {batch_file}")
    wb = load_workbook(batch_file, read_only=True, data_only=True)
    ws = wb.active

    headers = None
    orders  = {}

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            print(f"  Columns: {headers}")
            continue

        if not any(row):
            continue

        row_dict  = dict(zip(headers, row))
        order_num = str(row_dict.get("Order #", "") or "").strip()
        model_raw = str(row_dict.get("Model Number", "") or "").strip()
        model_key = model_raw.upper()
        truck     = str(row_dict.get("Truck", "") or "").strip().upper()

        if not order_num or order_num == "None":
            continue
        if truck in EXCLUDED_TRUCKS:
            continue

        if order_num not in orders:
            orders[order_num] = {"models": 0, "has_parts": False}

        if is_model(model_key):
            orders[order_num]["models"] += 1
        elif model_key not in DELETE_CODES:
            orders[order_num]["has_parts"] = True

    result = {order: max(1, data["models"]) + 1 for order, data in orders.items()}

    print(f"  Orders parsed: {len(result)}")
    for order, count in list(result.items())[:5]:
        print(f"    {order} -> {count} row(s)")

    return result


# ---------------------------------------------------------------------------
# Monday.com board operations
# ---------------------------------------------------------------------------

def get_board_items(token: str, board_id: str) -> tuple[list, list]:
    """Return (groups, all_items) for the given board, paginating as needed."""
    groups_query = """
    query ($boardId: [ID!]) {
      boards(ids: $boardId) {
        groups { id title }
      }
    }
    """
    r      = requests.post(MONDAY_API_URL, json={"query": groups_query, "variables": {"boardId": str(board_id)}}, headers=_monday_headers(token))
    groups = r.json()["data"]["boards"][0]["groups"]

    all_items = []
    cursor    = None

    while True:
        if cursor:
            q = """
            query ($boardId: [ID!], $cursor: String!) {
              boards(ids: $boardId) {
                items_page(limit: 500, cursor: $cursor) {
                  cursor
                  items { id name group { id } column_values { id value } }
                }
              }
            }
            """
            v = {"boardId": str(board_id), "cursor": cursor}
        else:
            q = """
            query ($boardId: [ID!]) {
              boards(ids: $boardId) {
                items_page(limit: 500) {
                  cursor
                  items { id name group { id } column_values { id value } }
                }
              }
            }
            """
            v = {"boardId": str(board_id)}

        r      = requests.post(MONDAY_API_URL, json={"query": q, "variables": v}, headers=_monday_headers(token))
        result = r.json()
        if "errors" in result:
            print(f"  get_board_items error: {result['errors']}")
            break
        page = result["data"]["boards"][0]["items_page"]
        all_items.extend(page["items"])
        cursor = page.get("cursor")
        if not cursor:
            break

    print(f"  Total items fetched: {len(all_items)}")
    return groups, all_items


def get_or_create_group(token: str, board_id: str, group_title: str, existing_groups: list) -> str:
    """Return the group ID for group_title, creating it if it doesn't exist."""
    for g in existing_groups:
        if g["title"] == group_title:
            print(f"  Group exists: {group_title}")
            return g["id"]

    mutation  = "mutation ($boardId: ID!, $groupName: String!) { create_group(board_id: $boardId, group_name: $groupName) { id } }"
    variables = {"boardId": str(board_id), "groupName": group_title}
    r         = requests.post(MONDAY_API_URL, json={"query": mutation, "variables": variables}, headers=_monday_headers(token))
    group_id  = r.json()["data"]["create_group"]["id"]
    print(f"  Created group: {group_title}")
    return group_id


def create_item(token: str, board_id: str, group_id: str, order_number: str, delivery_date: datetime):
    """Create a single Monday item for an order in the target group."""
    date_str      = delivery_date.strftime("%Y-%m-%d")
    column_values = json.dumps({"date4": {"date": date_str}})
    mutation      = "mutation ($boardId: ID!, $groupId: String!, $itemName: String!, $columnValues: JSON!) { create_item(board_id: $boardId, group_id: $groupId, item_name: $itemName, column_values: $columnValues) { id } }"
    variables     = {
        "boardId":      str(board_id),
        "groupId":      group_id,
        "itemName":     str(order_number),
        "columnValues": column_values,
    }
    r      = requests.post(MONDAY_API_URL, json={"query": mutation, "variables": variables}, headers=_monday_headers(token))
    result = r.json()
    if "errors" in result:
        print(f"  WARNING creating item {order_number}: {result['errors']}")


def populate_monday(token: str, board_id: str, orders: dict, delivery_date: datetime):
    """Create Monday items for each new order, skipping already-existing ones."""
    group_title = delivery_date.strftime("%m%d%y") + "PIX"
    today       = datetime.today().date()

    print("  Fetching board items...")
    existing_groups, all_items = get_board_items(token, board_id)

    skip_orders = set()
    for item in all_items:
        order_num = item["name"].strip()
        if item["group"]["id"] == TRANSFERRED_GROUP_ID:
            skip_orders.add(order_num)
            continue
        for col in item["column_values"]:
            if col["id"] == "date4" and col["value"]:
                try:
                    item_date = datetime.strptime(
                        json.loads(col["value"]).get("date", ""), "%Y-%m-%d"
                    ).date()
                    if item_date >= today:
                        skip_orders.add(order_num)
                except Exception:
                    pass
                break

    print(f"  Orders to skip: {len(skip_orders)}")
    group_id = get_or_create_group(token, board_id, group_title, existing_groups)

    created = skipped = 0
    for order_num, row_count in orders.items():
        if order_num in skip_orders:
            skipped += 1
            continue
        print(f"  Creating {row_count} row(s) for order {order_num}")
        for _ in range(row_count):
            create_item(token, board_id, group_id, order_num, delivery_date)
            time.sleep(0.3)
            created += 1

    print(f"  Done — created: {created}, skipped: {skipped}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run(hs_username: str, hs_password: str, monday_token: str, board_id: str) -> bool:
    delivery_date     = get_next_business_day()
    delivery_date_str = delivery_date.strftime("%B %d, %Y")

    print("\n=== MONDAY POPULATOR ===")
    print(f"Delivery date: {delivery_date_str}")

    batch_file = scrape_batch_invoice(hs_username, hs_password, delivery_date_str)
    if not batch_file:
        print("  ERROR: Could not download batch invoice")
        return False

    orders = parse_orders(batch_file)
    if not orders:
        print("  ERROR: No orders found")
        return False

    populate_monday(monday_token, board_id, orders, delivery_date)
    return True
