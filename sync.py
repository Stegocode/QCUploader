"""
sync.py — QCUploader Full Suite Orchestrator
=============================================
Runs every 30 minutes all day after the morning populator run.

Each cycle:
  1. Loads the batch invoice from DOWNLOAD_DIR (written by run_populator.py)
  2. Builds an order -> model map for mispick checking
  3. Fetches all items in the Monday completed group
  4. For each item:
       a. Runs mispick / misreceive check against the batch invoice
       b. Flagged items -> status updated in Monday, skipped
       c. Clean items -> photos downloaded and uploaded to platform
       d. Successful uploads -> marked as Transferred in Monday
  5. Logs results and sends failure email if needed
  6. Sleeps POLL_INTERVAL_MINS and repeats

Run with:
    python sync.py
"""

import asyncio
import glob
import os
import tempfile
import time
import traceback
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

import monday_client as mc
import mispick_checker as qc
from monday_populator import parse_orders, NON_MODEL_CODES, EXCLUDED_TRUCKS
from uploader import upload_order_photos

load_dotenv()

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MONDAY_TOKEN       = os.getenv("MONDAY_API_TOKEN")
BOARD_ID           = os.getenv("MONDAY_BOARD_ID")
HS_USERNAME        = os.getenv("HS_USERNAME")
HS_PASSWORD        = os.getenv("HS_PASSWORD")
POLL_INTERVAL_MINS = int(os.getenv("POLL_INTERVAL_MINS", "30"))

EMAIL_FROM     = os.getenv("EMAIL_FROM")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_TO       = os.getenv("EMAIL_TO", "")
SMTP_HOST      = os.getenv("SMTP_HOST", "smtp.office365.com")
SMTP_PORT      = int(os.getenv("SMTP_PORT", "587"))

APP_NAME     = os.getenv("APP_NAME", "QCUploader")
BASE_DIR     = Path(os.getenv("SYNC_BASE_DIR", Path.home() / "qcuploader"))
LOG_BASE     = BASE_DIR / "logs"
SCREEN_BASE  = BASE_DIR / "screenshots"
DOWNLOAD_DIR = Path(os.getenv("DOWNLOAD_DIR", Path.home() / "qcuploader" / "inbox"))


# ---------------------------------------------------------------------------
# Batch invoice loader
# ---------------------------------------------------------------------------

def find_batch_invoice() -> str | None:
    """Find the most recently downloaded batch invoice in DOWNLOAD_DIR."""
    matches = sorted(
        [f for f in glob.glob(str(DOWNLOAD_DIR / "*.xlsx")) if "bulk-invoice" in f.lower()],
        key=os.path.getmtime,
        reverse=True,
    )
    return matches[0] if matches else None


def load_order_model_map(batch_file: str) -> dict:
    """
    Parse the batch invoice and return a flat order -> model map.
    { "12345": "WFW5605MC", "12346": "MED6030MW", ... }
    Uses the same classification logic as the populator.
    Only the first/primary model per order is stored.
    """
    from openpyxl import load_workbook

    wb      = load_workbook(batch_file, read_only=True, data_only=True)
    ws      = wb.active
    headers = None
    result  = {}

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue

        row_dict  = dict(zip(headers, row))
        order_num = str(row_dict.get("Order #", "") or "").strip()
        model_raw = str(row_dict.get("Model Number", "") or "").strip()
        model_key = model_raw.upper()
        truck     = str(row_dict.get("Truck", "") or "").strip().upper()

        if not order_num or order_num == "None":
            continue
        if truck in EXCLUDED_TRUCKS:
            continue
        if model_key in NON_MODEL_CODES or not model_key:
            continue
        if order_num not in result:
            result[order_num] = model_key

    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def dated_folder(base: Path, dt: datetime) -> str:
    path = base / dt.strftime("%Y") / dt.strftime("%B").upper() / dt.strftime("%Y%m%d")
    path.mkdir(parents=True, exist_ok=True)
    return str(path)


def send_failure_email(failed_orders: list, results: dict, run_dt: datetime, err_log_path: str):
    """Send failure notification email when orders could not be processed."""
    if not EMAIL_FROM or not EMAIL_PASSWORD or not EMAIL_TO:
        print("  Email not configured — skipping notification")
        return

    recipients = [e.strip() for e in EMAIL_TO.split(",") if e.strip()]
    if not recipients:
        return

    run_str = run_dt.strftime("%B %d, %Y at %I:%M %p")
    subject = f"{APP_NAME} — {len(failed_orders)} order(s) failed {run_dt.strftime('%m/%d/%Y')}"

    lines = [f"The following orders failed during the {run_str} sync run:", ""]
    for order, reason in failed_orders:
        lines.append(f"  Order {order} — {reason}")

    lines += [
        "",
        "Run summary:",
        f"  SUCCESS    : {results.get('SUCCESS', 0)}",
        f"  FAILED     : {results.get('FAILED', 0)}",
        f"  MISPICK    : {results.get('MISPICK', 0)}",
        f"  MISRECEIVE : {results.get('MISRECEIVE', 0)}",
        f"  SKIPPED    : {results.get('SKIPPED', 0)}",
        "",
        "Please manually inspect flagged orders before next sync.",
        "",
        f"Error log  : {err_log_path}",
        f"Screenshots: {SCREEN_BASE}",
        "",
        f"— {APP_NAME}",
    ]

    try:
        msg = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = ", ".join(recipients)
        msg["Subject"] = subject
        msg.attach(MIMEText("\n".join(lines), "plain"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_FROM, EMAIL_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        print(f"  Email sent to: {', '.join(recipients)}")
    except Exception as e:
        print(f"  WARNING: Failed to send email: {e}")


# ---------------------------------------------------------------------------
# Run logger
# ---------------------------------------------------------------------------

class RunLogger:
    """Writes structured logs for a single sync cycle."""

    def __init__(self, run_dt: datetime):
        log_dir       = dated_folder(LOG_BASE, run_dt)
        ts            = run_dt.strftime("%Y%m%d_%H%M%S")
        self.run_path = os.path.join(log_dir, f"{ts}_run.log")
        self.err_path = os.path.join(log_dir, f"{ts}_errors.txt")
        self.run_dt   = run_dt
        self._errors  = []
        self._failed_orders = []

        self._append(
            self.run_path,
            "=" * 60 + f"\n  {APP_NAME}\n  Cycle started: {run_dt.strftime('%Y-%m-%d %H:%M:%S')}\n" + "=" * 60,
        )

    def _append(self, path: str, text: str):
        with open(path, "a", encoding="utf-8") as f:
            f.write(text + "\n")

    def _ts(self) -> str:
        return datetime.now().strftime("%H:%M:%S")

    def info(self, msg: str):
        line = f"[{self._ts()}] {msg}"
        print(line)
        self._append(self.run_path, line)

    def error(self, msg: str, exc=None):
        line = f"[{self._ts()}] ERROR: {msg}"
        print(line)
        self._append(self.run_path, line)
        entry = line
        if exc:
            tb = traceback.format_exc()
            entry += "\n" + tb
            self._append(self.run_path, tb)
        self._errors.append(entry)

    def result(self, order, status: str, reason: str = ""):
        line = f"[{self._ts()}] {status.upper().ljust(12)} | {order}"
        if reason:
            line += f" | {reason}"
        print(line)
        self._append(self.run_path, line)
        if status.upper() == "FAILED":
            self._failed_orders.append((str(order), reason or "unknown reason"))

    def finalize(self, results: dict):
        summary = (
            "\n" + "-" * 60 + "\n"
            f"  CYCLE SUMMARY\n"
            f"  SUCCESS    : {results.get('SUCCESS', 0)}\n"
            f"  FAILED     : {results.get('FAILED', 0)}\n"
            f"  MISPICK    : {results.get('MISPICK', 0)}\n"
            f"  MISRECEIVE : {results.get('MISRECEIVE', 0)}\n"
            f"  SKIPPED    : {results.get('SKIPPED', 0)}\n"
            + "-" * 60
        )
        self._append(self.run_path, summary)
        print(summary)

        if self._errors:
            self._append(
                self.err_path,
                f"ERRORS FROM CYCLE {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n" + "=" * 60,
            )
            for e in self._errors:
                self._append(self.err_path, e)
                self._append(self.err_path, "-" * 40)
            print(f"  Warning: {len(self._errors)} error(s) written to: {self.err_path}")
        else:
            print("  No errors this cycle.")

        if self._failed_orders:
            print("  Sending failure notification email...")
            send_failure_email(self._failed_orders, results, self.run_dt, self.err_path)


# ---------------------------------------------------------------------------
# Photo downloader
# ---------------------------------------------------------------------------

def download_photos_for_item(parsed_item: dict, logger: RunLogger) -> list:
    """Download serial and ID photos from Monday assets to a temp directory."""
    paths  = []
    tmpdir = tempfile.mkdtemp()
    order  = parsed_item["order_number"]

    for key, asset_id in [
        ("serial", parsed_item["serial_asset_id"]),
        ("id",     parsed_item["id_asset_id"]),
    ]:
        if not asset_id:
            continue
        url = mc.get_public_url(MONDAY_TOKEN, asset_id, logger)
        if not url:
            logger.error(f"Could not get public URL for {key} photo — order {order}")
            continue
        dest = os.path.join(tmpdir, f"{order}_{key}.jpg")
        if mc.download_photo(url, dest, logger):
            paths.append(dest)
        else:
            logger.error(f"Photo download failed — order {order} | {key}")

    return paths


# ---------------------------------------------------------------------------
# Main cycle
# ---------------------------------------------------------------------------

async def process_completed_items(run_dt: datetime, order_model_map: dict):
    logger     = RunLogger(run_dt)
    screen_dir = dated_folder(SCREEN_BASE, run_dt)
    results    = {"SUCCESS": 0, "FAILED": 0, "MISPICK": 0, "MISRECEIVE": 0, "SKIPPED": 0}

    logger.info("Checking completed items group...")
    items = mc.get_completed_items(MONDAY_TOKEN, BOARD_ID, logger)
    logger.info(f"Found {len(items)} item(s) to process")

    for item in items:
        order       = "UNKNOWN"
        photo_paths = []
        try:
            parsed    = mc.parse_item(item)
            order     = parsed["order_number"]
            logger.info(f"Processing order {order}...")

            # ── QC Gate ──────────────────────────────────────────────
            qc_parsed = qc.parse_item(item)
            qc_result = qc.check_item(qc_parsed, order_model_map)

            if qc_result["status"] == "NO_DATA":
                logger.result(order, "SKIPPED", "AI extraction not ready — retrying next cycle")
                results["SKIPPED"] += 1
                continue

            if qc_result["status"] in ("MISPICK", "BOTH"):
                logger.result(order, "MISPICK", qc_result["notes"])
                qc.update_status(MONDAY_TOKEN, BOARD_ID, parsed["item_id"], qc.STATUS_MISPICK, "MISPICK")
                results["MISPICK"] += 1
                continue

            if qc_result["status"] == "MISRECEIVE":
                logger.result(order, "MISRECEIVE", qc_result["notes"])
                qc.update_status(MONDAY_TOKEN, BOARD_ID, parsed["item_id"], qc.STATUS_MISRECEIVE, "MISRECEIVE")
                results["MISRECEIVE"] += 1
                continue

            if qc_result["status"] == "NO_MATCH":
                logger.error(f"Order {order} not in batch invoice — {qc_result['notes']}")
                logger.result(order, "SKIPPED", "not in batch invoice")
                results["SKIPPED"] += 1
                continue

            # ── QC passed — upload ────────────────────────────────────
            if not parsed["serial_asset_id"] and not parsed["id_asset_id"]:
                logger.result(order, "SKIPPED", "no photos attached")
                results["SKIPPED"] += 1
                continue

            photo_paths = download_photos_for_item(parsed, logger)
            if not photo_paths:
                logger.result(order, "FAILED", "photo download returned no files")
                results["FAILED"] += 1
                continue

            success = await upload_order_photos(
                HS_USERNAME, HS_PASSWORD, order, photo_paths,
                screenshot_dir=screen_dir,
            )

            if success:
                mc.mark_transferred(MONDAY_TOKEN, BOARD_ID, parsed["item_id"], logger)
                logger.result(order, "SUCCESS", f"model {qc_result['model_extract']}")
                results["SUCCESS"] += 1
            else:
                logger.result(order, "FAILED", "upload returned False")
                results["FAILED"] += 1

        except Exception as e:
            logger.error(f"Unhandled exception on order {order}: {e}", exc=e)
            logger.result(order, "FAILED", "unhandled exception — see error log")
            results["FAILED"] += 1
        finally:
            for path in photo_paths:
                try:
                    os.remove(path)
                except Exception:
                    pass

    logger.finalize(results)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run():
    print("=" * 60)
    print(f"  {APP_NAME} — Full Suite")
    print(f"  Polling every {POLL_INTERVAL_MINS} minutes")
    print(f"  Inbox      : {DOWNLOAD_DIR}")
    print(f"  Logs       : {LOG_BASE}")
    print(f"  Screenshots: {SCREEN_BASE}")
    print("  Press Ctrl+C to stop")
    print("=" * 60 + "\n")

    while True:
        run_dt = datetime.now()
        try:
            batch_file = find_batch_invoice()
            if not batch_file:
                print(f"  [{run_dt.strftime('%H:%M:%S')}] No batch invoice found in {DOWNLOAD_DIR}")
                print(f"  Run run_populator.py first to generate the batch invoice.")
            else:
                print(f"  [{run_dt.strftime('%H:%M:%S')}] Batch invoice: {Path(batch_file).name}")
                order_model_map = load_order_model_map(batch_file)
                print(f"  [{run_dt.strftime('%H:%M:%S')}] {len(order_model_map)} order(s) loaded")
                asyncio.run(process_completed_items(run_dt, order_model_map))

        except KeyboardInterrupt:
            print("\nStopped by user.")
            break
        except Exception as e:
            print(f"  Top-level run error: {e}")
            traceback.print_exc()

        print(f"\nSleeping {POLL_INTERVAL_MINS} minutes...\n")
        time.sleep(POLL_INTERVAL_MINS * 60)


if __name__ == "__main__":
    run()
